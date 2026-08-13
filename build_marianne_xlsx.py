import re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Pull Marianne's 16 records from maas-script.js ───────────────────────────
# We read line-by-line and grab each JSON object that has Michael Shonde as Rep
# AND Marianne Johnston as Tech Rep

lines = open('maas-script.js', encoding='utf-8').readlines()

# Find opening braces of each record that contains "Michael Shonde" as Rep
# Strategy: find each "Rep": "Michael Shonde" line, walk back to find the { ,
# walk forward to find the closing }, parse the block.

import json

def extract_block(lines, rep_line_idx):
    """Walk back to find '{' that opens this record, walk forward to closing '}'"""
    # walk back
    start = rep_line_idx
    while start > 0 and '{' not in lines[start]:
        start -= 1
    # walk forward to matching closing brace
    depth = 0
    end = start
    while end < len(lines):
        for ch in lines[end]:
            if ch == '{': depth += 1
            elif ch == '}': depth -= 1
        if depth == 0:
            break
        end += 1
    block = ''.join(lines[start:end+1]).strip().rstrip(',')
    return json.loads(block)

shonde_line_idxs = [i for i, l in enumerate(lines) if '"Rep": "Michael Shonde"' in l]

records = []
for idx in shonde_line_idxs:
    # Check it's one of Marianne's 16 (Tech Rep = Marianne Johnston)
    chunk = ''.join(lines[idx:idx+6])
    if 'Marianne Johnston' not in chunk:
        continue
    rec = extract_block(lines, idx)
    records.append(rec)

print(f'Found {len(records)} Marianne records')

# ── Build Excel ───────────────────────────────────────────────────────────────
IBM_BLUE  = '0F62FE'
WHITE     = 'FFFFFF'
YELLOW    = 'FFF8DC'
GREY_LOCK = 'E8E8E8'
TEXT_DARK = '161616'
MID_GREY  = 'E0E0E0'

HEADERS = [
    'COVERAGE_NAME', 'MARKET_NAME', 'SUB_MARKET_NAME', 'BRANCH_NAME', 'BRANCH_UNIT_NAME',
    'Rep', 'Role', 'Rep Email',
    'Tech Rep', 'Tech Rep Email',
    'Manager', 'Manager Email',
    'Tech Manager', 'Tech Manager Email',
]

# Columns that are read-only context (grey) vs editable (yellow)
LOCKED_COLS = {'COVERAGE_NAME','MARKET_NAME','SUB_MARKET_NAME','BRANCH_NAME','BRANCH_UNIT_NAME','Role'}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Marianne – Rep Alignment'

thin = Side(style='thin', color=MID_GREY)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Instructions row ──────────────────────────────────────────────────────────
ws.merge_cells(f'A1:{get_column_letter(len(HEADERS))}1')
ws['A1'] = (
    'Marianne — these are your 16 Rep Alignment records currently showing "Michael Shonde" as Rep. '
    'Please verify the yellow columns (Rep, Rep Email, Tech Rep, Tech Rep Email, Manager, etc.) '
    'and correct anything wrong. Send back to Bernie when done.'
)
ws['A1'].font = Font(name='Calibri', size=10, italic=True, color='525252')
ws['A1'].fill = PatternFill('solid', fgColor='F4F4F4')
ws['A1'].alignment = Alignment(wrap_text=True, vertical='center')
ws.row_dimensions[1].height = 36

# ── Header row ────────────────────────────────────────────────────────────────
for col_idx, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=2, column=col_idx, value=h)
    cell.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
    cell.fill = PatternFill('solid', fgColor=IBM_BLUE)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
ws.row_dimensions[2].height = 28

# ── Data rows ─────────────────────────────────────────────────────────────────
for row_idx, rec in enumerate(records, start=3):
    for col_idx, field in enumerate(HEADERS, start=1):
        value = rec.get(field, '')
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name='Calibri', size=10, color=TEXT_DARK)
        cell.alignment = Alignment(vertical='center', wrap_text=False)
        cell.border = border
        if field in LOCKED_COLS:
            cell.fill = PatternFill('solid', fgColor=GREY_LOCK)
            cell.font = Font(name='Calibri', size=10, color='525252')
        else:
            cell.fill = PatternFill('solid', fgColor=YELLOW)

# ── Column widths ─────────────────────────────────────────────────────────────
widths = [34, 22, 24, 26, 26, 20, 8, 28, 20, 28, 18, 28, 18, 28]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = 'A3'
ws.auto_filter.ref = f'A2:{get_column_letter(len(HEADERS))}2'

wb.save('marianne-rep-alignment-update.xlsx')
print('Saved → marianne-rep-alignment-update.xlsx')
