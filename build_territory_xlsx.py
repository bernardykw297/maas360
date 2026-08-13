import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Data (mirrors us-map-data.js exactly) ─────────────────────────────────────
STATES = [
    ('ME','Maine','James Chavez','James.Chavez@ibm.com','737-314-2537','Isabella Socarras','Isabella.Socarras@ibm.com','','Josh Knowles','jknowles@us.ibm.com','215-370-4706','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('NH','New Hampshire','James Chavez','James.Chavez@ibm.com','737-314-2537','Isabella Socarras','Isabella.Socarras@ibm.com','','Josh Knowles','jknowles@us.ibm.com','215-370-4706','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('VT','Vermont','James Chavez','James.Chavez@ibm.com','737-314-2537','Isabella Socarras','Isabella.Socarras@ibm.com','','Josh Knowles','jknowles@us.ibm.com','215-370-4706','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('MA','Massachusetts','James Chavez','James.Chavez@ibm.com','737-314-2537','Isabella Socarras','Isabella.Socarras@ibm.com','','Josh Knowles','jknowles@us.ibm.com','215-370-4706','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('RI','Rhode Island','James Chavez','James.Chavez@ibm.com','737-314-2537','Isabella Socarras','Isabella.Socarras@ibm.com','','Josh Knowles','jknowles@us.ibm.com','215-370-4706','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('CT','Connecticut','Asyah Jiron','a.jiron@ibm.com','','Isabella Socarras','Isabella.Socarras@ibm.com','','Marianne Johnston','mjohnston@ibm.com','215-237-4767','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('NJ','New Jersey','Asyah Jiron','a.jiron@ibm.com','','Isabella Socarras','Isabella.Socarras@ibm.com','','Marianne Johnston','mjohnston@ibm.com','215-237-4767','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('NY','New York','Steven St. Arnauld','steven.st.Arnauld@ibm.com','737-320-7342','Isabella Socarras','Isabella.Socarras@ibm.com','','Marianne Johnston','mjohnston@ibm.com','215-237-4767','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('PA','Pennsylvania','Robert Woessner','Robert.Woessner@ibm.com','737-615-9712','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Nate Alvin','nathan.alvis@ibm.com','469-887-3565','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('OH','Ohio','Robert Woessner','Robert.Woessner@ibm.com','737-615-9712','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Jason Sawyer + Fed','Jason.sawyer@ibm.com','512-799-9592','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('WV','West Virginia','Andrew Williams','Andrew.Williams.Jr@ibm.com','214-930-7274','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Nate Alvin','nathan.alvis@ibm.com','469-887-3565','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('VA','Virginia','Andrew Williams','Andrew.Williams.Jr@ibm.com','214-930-7274','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Nate Alvin','nathan.alvis@ibm.com','469-887-3565','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('MD','Maryland','Andrew Williams','Andrew.Williams.Jr@ibm.com','214-930-7274','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Nate Alvin','nathan.alvis@ibm.com','469-887-3565','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('DE','Delaware','Andrew Williams','Andrew.Williams.Jr@ibm.com','214-930-7274','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Nate Alvin','nathan.alvis@ibm.com','469-887-3565','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('KY','Kentucky','Nick Riley','Nicholas.Riley@ibm.com','817-691-6944','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Nathan Luu','nathan.luu@ibm.com','956-953-1495','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('TN','Tennessee','Nick Riley','Nicholas.Riley@ibm.com','817-691-6944','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Nathan Luu','nathan.luu@ibm.com','956-953-1495','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('NC','North Carolina','Nick Riley','Nicholas.Riley@ibm.com','817-691-6944','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Nathan Luu','nathan.luu@ibm.com','956-953-1495','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('MS','Mississippi','Sierra Reynolds','Sierra.Reynolds@ibm.com','949-282-7050','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Nathan Luu','nathan.luu@ibm.com','956-953-1495','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('AL','Alabama','Sierra Reynolds','Sierra.Reynolds@ibm.com','949-282-7050','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Nathan Luu','nathan.luu@ibm.com','956-953-1495','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('GA','Georgia','Sierra Reynolds','Sierra.Reynolds@ibm.com','949-282-7050','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Nathan Luu','nathan.luu@ibm.com','956-953-1495','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('SC','South Carolina','Sierra Reynolds','Sierra.Reynolds@ibm.com','949-282-7050','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Nathan Luu','nathan.luu@ibm.com','956-953-1495','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('FL','Florida',"Mary Kate O'Neill",'marykate.oneill@ibm.com','445-895-7443','Jake Neary','Jake.Neary@ibm.com','215-870-2905','Valery Orellana','valery.Orellana@ibm.com','512-673-8307','Mike Dailey','Senior Manager','Michael.Dailey1@ibm.com','215-983-4876','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Colby Smith','Colby.Smith@ibm.com','','Open','',''),
    ('MI','Michigan',"Bill D'Arcy",'Bill.DArcy@ibm.com','215-965-6271','Kory Kirby','Kory.Kirby@ibm.com','512-831-2699','Annika Harper','Annika.harper@ibm.com','737-587-7126','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('IN','Indiana',"Bill D'Arcy",'Bill.DArcy@ibm.com','215-965-6271','Kory Kirby','Kory.Kirby@ibm.com','512-831-2699','Annika Harper','Annika.harper@ibm.com','737-587-7126','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('IL','Illinois',"Bill D'Arcy",'Bill.DArcy@ibm.com','215-965-6271','Kory Kirby','Kory.Kirby@ibm.com','512-831-2699','Annika Harper','Annika.harper@ibm.com','737-587-7126','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('WI','Wisconsin',"Bill D'Arcy",'Bill.DArcy@ibm.com','215-965-6271','Kory Kirby','Kory.Kirby@ibm.com','512-831-2699','Annika Harper','Annika.harper@ibm.com','737-587-7126','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('MN','Minnesota','Erin Flynn','Erin.Flynn@ibm.com','215-883-6324','Kory Kirby','Kory.Kirby@ibm.com','512-831-2699','Alina Kim','alina.kim@ibm.com','737-376-4239','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('IA','Iowa','Erin Flynn','Erin.Flynn@ibm.com','215-883-6324','Kory Kirby','Kory.Kirby@ibm.com','512-831-2699','Alina Kim','alina.kim@ibm.com','737-376-4239','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('MO','Missouri','Erin Flynn','Erin.Flynn@ibm.com','215-883-6324','Kory Kirby','Kory.Kirby@ibm.com','512-831-2699','Alina Kim','alina.kim@ibm.com','737-376-4239','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('ND','North Dakota','Erin Flynn','Erin.Flynn@ibm.com','215-883-6324','Kory Kirby','Kory.Kirby@ibm.com','512-831-2699','Alina Kim','alina.kim@ibm.com','737-376-4239','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('SD','South Dakota','Erin Flynn','Erin.Flynn@ibm.com','215-883-6324','Kory Kirby','Kory.Kirby@ibm.com','512-831-2699','Alina Kim','alina.kim@ibm.com','737-376-4239','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('NE','Nebraska','Erin Flynn','Erin.Flynn@ibm.com','215-883-6324','Kory Kirby','Kory.Kirby@ibm.com','512-831-2699','Alina Kim','alina.kim@ibm.com','737-376-4239','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('KS','Kansas','Erin Flynn','Erin.Flynn@ibm.com','215-883-6324','Kory Kirby','Kory.Kirby@ibm.com','512-831-2699','Alina Kim','alina.kim@ibm.com','737-376-4239','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('AR','Arkansas','Erin Flynn','Erin.Flynn@ibm.com','215-883-6324','Nicole Matel','Nicole.Matel@ibm.com','','Alina Kim','alina.kim@ibm.com','737-376-4239','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('LA','Louisiana','Erin Flynn','Erin.Flynn@ibm.com','215-883-6324','Nicole Matel','Nicole.Matel@ibm.com','','Alina Kim','alina.kim@ibm.com','737-376-4239','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('OK','Oklahoma','Erin Flynn','Erin.Flynn@ibm.com','215-883-6324','Nicole Matel','Nicole.Matel@ibm.com','','Alina Kim','alina.kim@ibm.com','737-376-4239','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('TX','Texas','Jad Rusnol','Jad.Rusnol@ibm.com','726-248-3529','Nicole Matel','Nicole.Matel@ibm.com','','Duane Wright','wduane@us.ibm.com','215-528-3117','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('CO','Colorado','Allie Pitts','Allie.Pitts@ibm.com','214-707-3324','Nicole Matel','Nicole.Matel@ibm.com','','Brian Collier','brian.collier@ibm.com','717-826-3650','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('WY','Wyoming','Allie Pitts','Allie.Pitts@ibm.com','214-707-3324','Nicole Matel','Nicole.Matel@ibm.com','','Brian Collier','brian.collier@ibm.com','717-826-3650','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('MT','Montana','Allie Pitts','Allie.Pitts@ibm.com','214-707-3324','Nicole Matel','Nicole.Matel@ibm.com','','Brian Collier','brian.collier@ibm.com','717-826-3650','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('ID','Idaho','Allie Pitts','Allie.Pitts@ibm.com','214-707-3324','Nicole Matel','Nicole.Matel@ibm.com','','Brian Collier','brian.collier@ibm.com','717-826-3650','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('UT','Utah','Allie Pitts','Allie.Pitts@ibm.com','214-707-3324','Lesley Hazen','Lesley.Hazen@ibm.com','','Brian Collier','brian.collier@ibm.com','717-826-3650','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('NV','Nevada','Joe Clark','Joseph.Clark1@ibm.com','','Lesley Hazen','Lesley.Hazen@ibm.com','','Lina Corredor','lmcorred@us.ibm.com','727-810-6713','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('AZ','Arizona','Joe Clark','Joseph.Clark1@ibm.com','','Lesley Hazen','Lesley.Hazen@ibm.com','','Lina Corredor','lmcorred@us.ibm.com','727-810-6713','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('NM','New Mexico','Joe Clark','Joseph.Clark1@ibm.com','','Lesley Hazen','Lesley.Hazen@ibm.com','','Lina Corredor','lmcorred@us.ibm.com','727-810-6713','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('CA','California','Zach Tarver','Zach.Tarver@ibm.com','945-324-0588','Lesley Hazen','Lesley.Hazen@ibm.com','','Ruby Ramsay','rubyramsay@ibm.com','512-937-5814','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('OR','Oregon','Allie Pitts','Allie.Pitts@ibm.com','214-707-3324','Lesley Hazen','Lesley.Hazen@ibm.com','','Brian Collier','brian.collier@ibm.com','717-826-3650','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('WA','Washington','Allie Pitts','Allie.Pitts@ibm.com','214-707-3324','Lesley Hazen','Lesley.Hazen@ibm.com','','Brian Collier','brian.collier@ibm.com','717-826-3650','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('AK','Alaska','Joe Clark','Joseph.Clark1@ibm.com','','Lesley Hazen','Lesley.Hazen@ibm.com','','Brian Collier','brian.collier@ibm.com','717-826-3650','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
    ('HI','Hawaii','Joe Clark','Joseph.Clark1@ibm.com','','Lesley Hazen','Lesley.Hazen@ibm.com','','Brian Collier','brian.collier@ibm.com','717-826-3650','Kellin Johnson','Senior Manager','kellinj@us.ibm.com','972-841-0826','Colin Shenk','Colin.Shenk@ibm.com','(717) 368-2766','Hemali Patel','H.Patel@ibm.com','828-507-6563','Eric Szewczyk','eric.s@ibm.com','512-987-4769'),
]

HEADERS = [
    'State Code', 'State Name',
    'Rep', 'Rep Email', 'Rep Phone',
    'Telco Specialist', 'Telco Email', 'Telco Phone',
    'Tech Rep', 'Tech Rep Email', 'Tech Rep Phone',
    'Manager', 'Manager Title', 'Manager Email', 'Manager Phone',
    'Tech Manager', 'Tech Manager Email', 'Tech Manager Phone',
    'AT&T Specialist', 'AT&T Email', 'AT&T Phone',
    'T-Mobile Specialist', 'T-Mobile Email', 'T-Mobile Phone',
]

# ── Colour palette (IBM Carbon-ish) ───────────────────────────────────────────
IBM_BLUE      = '0F62FE'
IBM_BLUE_DARK = '0043CE'
WHITE         = 'FFFFFF'
LIGHT_GREY    = 'F4F4F4'
MID_GREY      = 'E0E0E0'
TEXT_DARK     = '161616'

# Column group fills
GROUP_FILLS = {
    'rep':     'EDF5FF',   # light blue
    'telco':   'DEFBE6',   # light green
    'tech':    'FFF1F1',   # light red/pink
    'manager': 'F6F2FF',   # light purple
    'techman': 'FEF6E4',   # light amber
    'att':     'F6FEFE',   # light teal
    'tmobile': 'FFF0F7',   # light magenta
}
# Which column indices (0-based, after State Code/Name) belong to each group
COL_GROUPS = [
    ('rep',     [2,3,4]),
    ('telco',   [5,6,7]),
    ('tech',    [8,9,10]),
    ('manager', [11,12,13,14]),
    ('techman', [15,16,17]),
    ('att',     [18,19,20]),
    ('tmobile', [21,22,23]),
]

def col_fill(col_idx):
    for name, indices in COL_GROUPS:
        if col_idx in indices:
            return GROUP_FILLS[name]
    return LIGHT_GREY  # State Code / State Name

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Territory Assignments'

# ── Instructions row ─────────────────────────────────────────────────────────
ws.merge_cells('A1:X1')
ws['A1'] = (
    'MaaS360 Territory Assignments — Fill in / correct your rows, then send back to Bernie. '
    'Yellow cells = editable fields. Do not change column headers or State Code / State Name columns.'
)
ws['A1'].font = Font(name='Calibri', size=10, italic=True, color='525252')
ws['A1'].fill = PatternFill('solid', fgColor='F4F4F4')
ws['A1'].alignment = Alignment(wrap_text=True, vertical='center')
ws.row_dimensions[1].height = 30

# ── Group label row ───────────────────────────────────────────────────────────
group_labels = [
    (1,  2,  ''),
    (3,  5,  'REP'),
    (6,  8,  'TELCO SPECIALIST'),
    (9,  11, 'TECH REP'),
    (12, 15, 'MANAGER'),
    (16, 18, 'TECH MANAGER'),
    (19, 21, 'AT&T SPECIALIST'),
    (22, 24, 'T-MOBILE SPECIALIST'),
]
for start_col, end_col, label in group_labels:
    if start_col == end_col:
        cell = ws.cell(row=2, column=start_col, value=label)
    else:
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        cell = ws.cell(row=2, column=start_col, value=label)
    fill_key = None
    for name, indices in COL_GROUPS:
        if (start_col - 1) in indices:
            fill_key = name
            break
    bg = GROUP_FILLS.get(fill_key, IBM_BLUE_DARK) if fill_key else IBM_BLUE_DARK
    cell.fill = PatternFill('solid', fgColor=bg if fill_key else IBM_BLUE_DARK)
    cell.font = Font(name='Calibri', size=9, bold=True,
                     color=TEXT_DARK if fill_key else WHITE)
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 18

# ── Header row ────────────────────────────────────────────────────────────────
header_fill = PatternFill('solid', fgColor=IBM_BLUE)
for col_idx, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[3].height = 28

# ── Data rows ─────────────────────────────────────────────────────────────────
thin = Side(style='thin', color=MID_GREY)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

editable_fill   = PatternFill('solid', fgColor='FFFFF0')   # very light yellow — editable
readonly_fill_e = PatternFill('solid', fgColor='F4F4F4')   # grey — state code/name

for row_idx, state_row in enumerate(STATES, start=4):
    is_even = (row_idx % 2 == 0)
    for col_idx, value in enumerate(state_row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name='Calibri', size=10, color=TEXT_DARK)
        cell.alignment = Alignment(vertical='center')
        cell.border = border
        if col_idx <= 2:
            # State code / name — lock visually
            cell.fill = PatternFill('solid', fgColor='E8E8E8' if is_even else 'F0F0F0')
            cell.font = Font(name='Calibri', size=10, bold=(col_idx==1), color='393939')
        else:
            # Editable — light group colour
            bg = col_fill(col_idx - 1)
            # Slightly darker on even rows for readability
            cell.fill = PatternFill('solid', fgColor=bg)

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = [7, 16, 20, 28, 14, 20, 28, 14, 20, 28, 14, 18, 16, 28, 14, 18, 28, 14, 18, 28, 14, 20, 28, 14]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Freeze panes (keep header + state columns visible) ────────────────────────
ws.freeze_panes = 'C4'

# ── Auto-filter on header row ─────────────────────────────────────────────────
ws.auto_filter.ref = f'A3:{get_column_letter(len(HEADERS))}3'

# ── Instructions sheet ───────────────────────────────────────────────────────
ws2 = wb.create_sheet('How to Use')
instructions = [
    ('MaaS360 Territory Assignment Update', True, 13),
    ('', False, 11),
    ('STEP 1 — Find your states', True, 11),
    ('  Use the filter arrows on row 3 (the header row) to filter by Rep, Tech Rep, or any other column.', False, 10),
    ('  Or just scroll — states are in geographic order from east to west.', False, 10),
    ('', False, 10),
    ('STEP 2 — Edit your cells', True, 11),
    ('  Click any yellow/coloured cell and type your correction.', False, 10),
    ('  Do NOT change column A (State Code) or column B (State Name).', False, 10),
    ('  Do NOT add or remove rows.', False, 10),
    ('', False, 10),
    ('STEP 3 — Send it back', True, 11),
    ('  Save the file and email it back to Bernie.', False, 10),
    ('  Bernie will apply all changes to the live Territory Map in one go.', False, 10),
    ('', False, 10),
    ('COLUMN GUIDE', True, 11),
    ('  Rep / Rep Email / Rep Phone          → The primary sales rep for this state', False, 10),
    ('  Telco Specialist / Email / Phone     → Telco overlay specialist', False, 10),
    ('  Tech Rep / Email / Phone             → Technical sales rep (SE)', False, 10),
    ('  Manager / Title / Email / Phone      → Sales manager', False, 10),
    ('  Tech Manager / Email / Phone         → Technical manager', False, 10),
    ('  AT&T Specialist / Email / Phone      → AT&T carrier specialist', False, 10),
    ('  T-Mobile Specialist / Email / Phone  → T-Mobile carrier specialist', False, 10),
]
for r, (text, bold, size) in enumerate(instructions, start=1):
    cell = ws2.cell(row=r, column=1, value=text)
    cell.font = Font(name='Calibri', size=size, bold=bold, color=TEXT_DARK)
ws2.column_dimensions['A'].width = 80

wb.save('territory-assignments.xlsx')
print('Done — territory-assignments.xlsx created.')
